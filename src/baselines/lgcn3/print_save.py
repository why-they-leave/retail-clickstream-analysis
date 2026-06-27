import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
import os

def print_params(para_name, para):
    for i in range(len(para)):
        print(para_name[i]+':  ',para[i])

def print_value(value):
    [inter, loss, f1_max, F1, NDCG] = value
    print('iter: %d loss %.2f f1 %.4f' %(inter, loss, f1_max), end='  ')
    print(F1, NDCG)

def save_params(para_name,para,path_excel):
    wb = Workbook( )
    table = wb.active
    table.title = 'Parameters'
    ldata = []
    for i in range(1, len(para_name)):  # do not save GPU_index
        parameter = [para_name[i]]
        parameter_value = para[i]
        if isinstance(parameter_value, list):
            for value in parameter_value:
                parameter.append(value)
        elif isinstance(parameter_value, bool): parameter.append({True: 'Yes', False: 'No'}[parameter_value])
        else: parameter.append(parameter_value)
        ldata.append(parameter)
    for i, p in enumerate(ldata):
        for j, q in enumerate(p):
            table.cell(row = i+1, column = j+1).value = q
    wb.save(path_excel)
    # wb.close()

def save_value(df_list, path_excel, first_sheet=False):
    # df_list: [[F1_df, 'F1'], [NDCG_df, 'NDCG']], 
    if not os.path.exists(path_excel):
        df = pd.DataFrame()
        df.to_excel(path_excel)
    with pd.ExcelWriter(path_excel, mode='a', engine="openpyxl", if_sheet_exists="replace",) as excelWriter:
        for df, sheet_name in df_list:
            df.to_excel(excel_writer=excelWriter, sheet_name=sheet_name, index=True)

def df2str(df):
    df_str = ''
    for i in range(df.shape[0]):
        df_list = df.iloc[[i], :].values.tolist()
        df_list2 = [str(i) for i in df_list]
        str_temp = ''.join(df_list2)
        df_str = df_str +str_temp+','
    return df_str